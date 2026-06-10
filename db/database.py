"""
Database module for Prometheus.
Handles SQLite connection, schema creation, and all database operations.
"""

import sqlite3
import shutil
import os
from datetime import datetime, date
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from contextlib import contextmanager

from app.config import config


# --- Connection Management ---

@contextmanager
def get_connection():
    """Context manager for database connections."""
    conn = sqlite3.connect(config.DATABASE_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute("PRAGMA journal_mode = WAL")   # better concurrency
    try:
        yield conn
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# --- Schema ---

INIT_SQL = """
-- Items table (main inventory)
CREATE TABLE IF NOT EXISTS items (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    item_id TEXT UNIQUE NOT NULL,
    item_name TEXT NOT NULL,
    unit TEXT NOT NULL DEFAULT 'pcs',
    starting_stock REAL NOT NULL DEFAULT 0,
    current_stock REAL NOT NULL DEFAULT 0,
    used REAL NOT NULL DEFAULT 0,
    purpose TEXT DEFAULT '',
    wipro_in REAL NOT NULL DEFAULT 0,
    wipro_out REAL NOT NULL DEFAULT 0,
    rajagiri_main REAL NOT NULL DEFAULT 0,
    woods REAL NOT NULL DEFAULT 0,
    garden_cafe REAL NOT NULL DEFAULT 0,
    bba_canteen REAL NOT NULL DEFAULT 0,
    bba_tea_counter REAL NOT NULL DEFAULT 0,
    purchased REAL NOT NULL DEFAULT 0,
    damaged REAL NOT NULL DEFAULT 0,
    avg_daily_usage REAL NOT NULL DEFAULT 0,
    location TEXT DEFAULT '',
    category TEXT DEFAULT '',
    last_updated TEXT,
    last_updated_by TEXT DEFAULT '',
    working_date TEXT
);

-- Transactions table (audit log)
CREATE TABLE IF NOT EXISTS transactions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT NOT NULL,
    date TEXT NOT NULL,
    item_id TEXT NOT NULL,
    item_name TEXT NOT NULL,
    action TEXT NOT NULL,
    quantity REAL NOT NULL,
    unit TEXT NOT NULL,
    stock_before REAL NOT NULL,
    stock_after REAL NOT NULL,
    purpose TEXT DEFAULT '',
    destination TEXT DEFAULT '',
    performed_by TEXT DEFAULT '',
    notes TEXT DEFAULT ''
);

-- Purposes table
CREATE TABLE IF NOT EXISTS purposes (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT UNIQUE NOT NULL
);

-- Settings table
CREATE TABLE IF NOT EXISTS settings (
    key TEXT PRIMARY KEY,
    value TEXT NOT NULL
);

-- Backups table (metadata)
CREATE TABLE IF NOT EXISTS backups (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    filename TEXT NOT NULL,
    created_at TEXT NOT NULL,
    note TEXT DEFAULT ''
);

-- Undo log table
CREATE TABLE IF NOT EXISTS undo_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT NOT NULL,
    transaction_id INTEGER,
    item_id TEXT NOT NULL,
    action TEXT NOT NULL,
    quantity REAL NOT NULL,
    stock_before REAL NOT NULL,
    stock_after REAL NOT NULL,
    details TEXT DEFAULT '',
    performed_by TEXT DEFAULT '',
    reversed INTEGER DEFAULT 0
);

-- Indexes for performance
CREATE INDEX IF NOT EXISTS idx_items_name ON items(item_name);
CREATE INDEX IF NOT EXISTS idx_items_category ON items(category);
CREATE INDEX IF NOT EXISTS idx_transactions_date ON transactions(date);
CREATE INDEX IF NOT EXISTS idx_transactions_item ON transactions(item_id);
CREATE INDEX IF NOT EXISTS idx_undo_log_timestamp ON undo_log(timestamp);
"""


def init_db() -> None:
    """Initialize the database with all tables."""
    with get_connection() as conn:
        conn.executescript(INIT_SQL)
    ensure_default_purposes()
    ensure_settings()


def ensure_default_purposes() -> None:
    """Insert default purposes if the table is empty."""
    defaults = [
        "biriyani", "meals", "sambar", "breakfast", "snacks",
        "special event", "regular cooking", "trial", "waste",
    ]
    with get_connection() as conn:
        cursor = conn.execute("SELECT COUNT(*) as count FROM purposes")
        count = cursor.fetchone()["count"]
        if count == 0:
            for purpose in defaults:
                conn.execute(
                    "INSERT OR IGNORE INTO purposes (name) VALUES (?)",
                    (purpose,)
                )


def ensure_settings() -> None:
    """Ensure default settings exist (INSERT OR IGNORE keeps existing values)."""
    defaults = {
        "secret_word": config.SECRET_WORD,
        "reorder_days": str(config.REORDER_DAYS),
        "working_date": date.today().isoformat(),
    }
    with get_connection() as conn:
        for key, value in defaults.items():
            conn.execute(
                "INSERT OR IGNORE INTO settings (key, value) VALUES (?, ?)",
                (key, value)
            )


# --- Item CRUD ---

def add_item(
    item_id: str,
    item_name: str,
    unit: str = "pcs",
    starting_stock: float = 0,
    current_stock: float = 0,
    avg_daily_usage: float = 0,
    location: str = "",
    category: str = "",
    working_date: str = None,
) -> Tuple[bool, str]:
    """Add a new item to inventory."""
    if working_date is None:
        working_date = date.today().isoformat()
    try:
        with get_connection() as conn:
            conn.execute(
                """INSERT INTO items (
                    item_id, item_name, unit, starting_stock, current_stock,
                    avg_daily_usage, location, category, working_date, last_updated
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    item_id, item_name, unit, starting_stock, current_stock,
                    avg_daily_usage, location, category, working_date,
                    datetime.now().isoformat()
                )
            )
        return True, f"Item '{item_name}' added successfully."
    except sqlite3.IntegrityError:
        return False, f"Item ID '{item_id}' already exists."
    except Exception as e:
        return False, f"Error adding item: {str(e)}"


def get_item_by_id(item_id: str) -> Optional[Dict[str, Any]]:
    """Get a single item by its item_id."""
    with get_connection() as conn:
        cursor = conn.execute("SELECT * FROM items WHERE item_id = ?", (item_id,))
        row = cursor.fetchone()
        return dict(row) if row else None


def get_item_by_name(item_name: str) -> Optional[Dict[str, Any]]:
    """Get a single item by its name (case-insensitive, exact match)."""
    with get_connection() as conn:
        cursor = conn.execute(
            "SELECT * FROM items WHERE LOWER(item_name) = LOWER(?)",
            (item_name,)
        )
        row = cursor.fetchone()
        return dict(row) if row else None


def find_items_by_name_partial(name_partial: str) -> List[Dict[str, Any]]:
    """Find items by partial name match (case-insensitive)."""
    with get_connection() as conn:
        cursor = conn.execute(
            "SELECT * FROM items WHERE LOWER(item_name) LIKE LOWER(?) ORDER BY item_name",
            (f"%{name_partial}%",)
        )
        return [dict(row) for row in cursor.fetchall()]


def list_all_items() -> List[Dict[str, Any]]:
    """List all inventory items ordered by name."""
    with get_connection() as conn:
        cursor = conn.execute("SELECT * FROM items ORDER BY item_name")
        return [dict(row) for row in cursor.fetchall()]


def list_items_by_category(category: str) -> List[Dict[str, Any]]:
    """List items filtered by category."""
    with get_connection() as conn:
        cursor = conn.execute(
            "SELECT * FROM items WHERE LOWER(category) = LOWER(?) ORDER BY item_name",
            (category,)
        )
        return [dict(row) for row in cursor.fetchall()]


# Whitelist of valid section column names (prevents SQL injection)
_SECTION_COLUMN_MAP: Dict[str, str] = {
    "rajagiri main":   "rajagiri_main",
    "woods":           "woods",
    "garden cafe":     "garden_cafe",
    "bba canteen":     "bba_canteen",
    "bba tea counter": "bba_tea_counter",
    "wipro in":        "wipro_in",
    "wipro out":       "wipro_out",
}


def list_items_by_section(section: str) -> List[Dict[str, Any]]:
    """
    List items that had non-zero movement for a given section today.
    Only whitelisted column names are allowed (no SQL injection possible).
    """
    col = _SECTION_COLUMN_MAP.get(section.lower())
    if not col:
        return []   # unknown section -> return empty safely
    with get_connection() as conn:
        cursor = conn.execute(
            f"SELECT * FROM items WHERE {col} > 0 ORDER BY item_name"
        )
        return [dict(row) for row in cursor.fetchall()]


def update_item(
    item_id: str,
    updates: Dict[str, Any],
    performed_by: str = ""
) -> Tuple[bool, str]:
    """Update item fields. Returns (success, message)."""
    allowed_fields = {
        "item_name", "unit", "starting_stock", "current_stock",
        "avg_daily_usage", "location", "category", "purpose",
        "wipro_in", "wipro_out", "rajagiri_main", "woods",
        "garden_cafe", "bba_canteen", "bba_tea_counter",
        "purchased", "damaged", "used",
    }
    filtered = {k: v for k, v in updates.items() if k in allowed_fields}
    if not filtered:
        return False, "No valid fields to update."

    filtered["last_updated"] = datetime.now().isoformat()
    if performed_by:
        filtered["last_updated_by"] = performed_by

    set_clause = ", ".join(f"{k} = ?" for k in filtered.keys())
    values = list(filtered.values()) + [item_id]

    try:
        with get_connection() as conn:
            cursor = conn.execute(
                f"UPDATE items SET {set_clause} WHERE item_id = ?",
                values
            )
            if cursor.rowcount == 0:
                return False, f"Item '{item_id}' not found."
        return True, f"Item '{item_id}' updated successfully."
    except Exception as e:
        return False, f"Error updating item: {str(e)}"


def delete_item(item_id: str) -> Tuple[bool, str]:
    """Delete a single item."""
    with get_connection() as conn:
        cursor = conn.execute("DELETE FROM items WHERE item_id = ?", (item_id,))
        if cursor.rowcount == 0:
            return False, f"Item '{item_id}' not found."
    return True, f"Item '{item_id}' deleted successfully."


def delete_all_items() -> Tuple[bool, str]:
    """Delete all items and their related records."""
    with get_connection() as conn:
        conn.execute("DELETE FROM items")
        conn.execute("DELETE FROM transactions")
        conn.execute("DELETE FROM undo_log")
    return True, "All items and related records have been deleted."


# --- Stock Movements ---

def record_transaction(
    conn: sqlite3.Connection,
    item_id: str,
    item_name: str,
    action: str,
    quantity: float,
    unit: str,
    stock_before: float,
    stock_after: float,
    purpose: str = "",
    destination: str = "",
    performed_by: str = "",
    notes: str = "",
) -> int:
    """Record a transaction and return its ID."""
    today = date.today().isoformat()
    cursor = conn.execute(
        """INSERT INTO transactions (
            timestamp, date, item_id, item_name, action, quantity, unit,
            stock_before, stock_after, purpose, destination, performed_by, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            datetime.now().isoformat(), today, item_id, item_name,
            action, quantity, unit, stock_before, stock_after,
            purpose, destination, performed_by, notes
        )
    )
    return cursor.lastrowid


def log_undo_action(
    conn: sqlite3.Connection,
    transaction_id: int,
    item_id: str,
    action: str,
    quantity: float,
    stock_before: float,
    stock_after: float,
    details: str = "",
    performed_by: str = "",
) -> None:
    """Log an action to the undo log."""
    conn.execute(
        """INSERT INTO undo_log (
            timestamp, transaction_id, item_id, action, quantity,
            stock_before, stock_after, details, performed_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            datetime.now().isoformat(), transaction_id, item_id, action,
            quantity, stock_before, stock_after, details, performed_by
        )
    )


# Maps each action to its sub-column in the items table
_ACTION_COLUMN_MAP: Dict[str, str] = {
    "used":            "used",
    "purchased":       "purchased",
    "damaged":         "damaged",
    "wipro in":        "wipro_in",
    "wipro out":       "wipro_out",
    "rajagiri main":   "rajagiri_main",
    "woods":           "woods",
    "garden cafe":     "garden_cafe",
    "bba canteen":     "bba_canteen",
    "bba tea counter": "bba_tea_counter",
}

# Actions that reduce stock
_SUBTRACT_ACTIONS = frozenset({
    "used", "damaged", "wipro out",
    "rajagiri main", "woods", "garden cafe",
    "bba canteen", "bba tea counter",
})

# Actions that increase stock
_ADD_ACTIONS = frozenset({"purchased", "wipro in"})

# Transfer actions - allowed to go below zero
_TRANSFER_ACTIONS = frozenset({
    "wipro out", "rajagiri main", "woods",
    "garden cafe", "bba canteen", "bba tea counter",
})


def perform_stock_movement(
    item_id: str,
    action: str,
    quantity: float,
    purpose: str = "",
    destination: str = "",
    performed_by: str = "",
    notes: str = "",
) -> Tuple[bool, str, Optional[int]]:
    """
    Perform a stock movement and return (success, message, transaction_id).
    """
    item = get_item_by_id(item_id)
    if not item:
        return False, f"Item '{item_id}' not found.", None

    stock_before = item["current_stock"]
    unit = item["unit"]
    item_name = item["item_name"]

    column_map: Dict[str, Tuple[str, str]] = {
        "used":            ("used",            "subtract"),
        "purchased":       ("purchased",       "add"),
        "damaged":         ("damaged",         "subtract"),
        "wipro in":        ("wipro_in",        "add"),
        "wipro out":       ("wipro_out",       "subtract"),
        "rajagiri main":   ("rajagiri_main",   "subtract"),
        "woods":           ("woods",           "subtract"),
        "garden cafe":     ("garden_cafe",     "subtract"),
        "bba canteen":     ("bba_canteen",     "subtract"),
        "bba tea counter": ("bba_tea_counter", "subtract"),
    }

    try:
        with get_connection() as conn:
            # -- Stock adjustment: set absolute value --
            if action == "stock adjustment":
                new_stock = quantity
                adjustment = quantity - stock_before
                conn.execute(
                    "UPDATE items SET current_stock = ?, last_updated = ? WHERE item_id = ?",
                    (new_stock, datetime.now().isoformat(), item_id)
                )
                tx_id = record_transaction(
                    conn, item_id, item_name, action, abs(adjustment), unit,
                    stock_before, new_stock, performed_by=performed_by,
                    notes=f"Adjusted from {stock_before} to {new_stock}"
                )
                log_undo_action(
                    conn, tx_id, item_id, action, adjustment,
                    stock_before, new_stock, performed_by=performed_by
                )
                return (
                    True,
                    f"Stock adjusted: {item_name} from {stock_before} to {new_stock} {unit}.",
                    tx_id,
                )

            # -- Direct field edits (no undo logged - admin action) --
            elif action in ("change starting stock", "change current stock"):
                field = "starting_stock" if action == "change starting stock" else "current_stock"
                conn.execute(
                    f"UPDATE items SET {field} = ?, last_updated = ? WHERE item_id = ?",
                    (quantity, datetime.now().isoformat(), item_id)
                )
                tx_id = record_transaction(
                    conn, item_id, item_name, action, 0, unit,
                    stock_before, quantity, performed_by=performed_by,
                    notes=f"{action}: set to {quantity}"
                )
                return (
                    True,
                    f"{action.title()}: {item_name} set to {quantity} {unit}.",
                    tx_id,
                )

            # -- Normal movement --
            elif action in column_map:
                col, operation = column_map[action]

                if operation == "subtract":
                    new_stock = stock_before - quantity
                    # Only block negative stock for consumable actions (used/damaged)
                    if new_stock < 0 and action not in _TRANSFER_ACTIONS:
                        return (
                            False,
                            f"Insufficient stock. Current: {stock_before} {unit},"
                            f" requested: {quantity} {unit}.",
                            None,
                        )
                else:
                    new_stock = stock_before + quantity

                conn.execute(
                    f"UPDATE items SET {col} = {col} + ?, current_stock = ?,"
                    f" last_updated = ? WHERE item_id = ?",
                    (quantity, new_stock, datetime.now().isoformat(), item_id)
                )
                tx_id = record_transaction(
                    conn, item_id, item_name, action, quantity, unit,
                    stock_before, new_stock, purpose=purpose,
                    destination=destination, performed_by=performed_by, notes=notes
                )
                log_undo_action(
                    conn, tx_id, item_id, action, quantity,
                    stock_before, new_stock,
                    details=purpose or destination or "",
                    performed_by=performed_by
                )
                return (
                    True,
                    f"{action.title()}: {item_name} {quantity} {unit}."
                    f" Stock: {stock_before} -> {new_stock}.",
                    tx_id,
                )

            else:
                return False, f"Unknown action: '{action}'.", None

    except Exception as e:
        return False, f"Error performing {action}: {str(e)}", None


# --- Queries & Reports ---

def get_low_stock_items() -> List[Dict[str, Any]]:
    """Return items whose days-of-stock remaining is <= reorder threshold."""
    with get_connection() as conn:
        items = [dict(row) for row in conn.execute("SELECT * FROM items").fetchall()]

    reorder_days = config.REORDER_DAYS
    low_stock = []
    for item in items:
        avg = item["avg_daily_usage"] or 0
        current = item["current_stock"] or 0
        if avg > 0:
            days_left = current / avg
        else:
            days_left = 9999  # no usage data -> never triggers reorder

        if current <= 0 or days_left <= reorder_days:
            item["days_left"] = round(days_left, 1) if avg > 0 else "N/A"
            item["urgent"] = current <= 0
            low_stock.append(item)

    return low_stock


def get_order_list() -> List[Dict[str, Any]]:
    """Alias for get_low_stock_items (used for order list display)."""
    return get_low_stock_items()


def get_zero_stock_items() -> List[Dict[str, Any]]:
    """Return items with zero or negative stock."""
    with get_connection() as conn:
        cursor = conn.execute(
            "SELECT * FROM items WHERE current_stock <= 0 ORDER BY item_name"
        )
        return [dict(row) for row in cursor.fetchall()]


def get_daily_report() -> Dict[str, Any]:
    """Build today's full activity report."""
    today = date.today().isoformat()
    with get_connection() as conn:
        transactions = [
            dict(row)
            for row in conn.execute(
                "SELECT * FROM transactions WHERE date = ? ORDER BY timestamp",
                (today,)
            ).fetchall()
        ]

    def _filter(action_val):
        return [t for t in transactions if t["action"] == action_val]

    def _filter_in(action_vals):
        return [t for t in transactions if t["action"] in action_vals]

    return {
        "date": today,
        "used_items":      _filter("used"),
        "purchased_items": _filter("purchased"),
        "damaged_items":   _filter("damaged"),
        "wipro_in":        _filter("wipro in"),
        "wipro_out":       _filter("wipro out"),
        "transfers":       _filter_in({
            "rajagiri main", "woods", "garden cafe", "bba canteen", "bba tea counter"
        }),
        "adjustments":     _filter("stock adjustment"),
        "edits":           _filter_in({"change starting stock", "change current stock"}),
        "stock_status":    list_all_items(),
        "low_stock":       get_low_stock_items(),
        "total_transactions": len(transactions),
    }


# --- Purposes ---

def list_purposes() -> List[str]:
    with get_connection() as conn:
        return [row["name"] for row in conn.execute("SELECT name FROM purposes ORDER BY name")]


def add_purpose(name: str) -> Tuple[bool, str]:
    try:
        with get_connection() as conn:
            conn.execute("INSERT INTO purposes (name) VALUES (?)", (name,))
        return True, f"Purpose '{name}' added."
    except sqlite3.IntegrityError:
        return False, f"Purpose '{name}' already exists."


def remove_purpose(name: str) -> Tuple[bool, str]:
    with get_connection() as conn:
        cursor = conn.execute("DELETE FROM purposes WHERE name = ?", (name,))
        if cursor.rowcount == 0:
            return False, f"Purpose '{name}' not found."
    return True, f"Purpose '{name}' removed."


# --- Settings ---

def get_setting(key: str, default: str = "") -> str:
    with get_connection() as conn:
        cursor = conn.execute("SELECT value FROM settings WHERE key = ?", (key,))
        row = cursor.fetchone()
        return row["value"] if row else default


def set_setting(key: str, value: str) -> None:
    with get_connection() as conn:
        conn.execute(
            "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
            (key, value)
        )


def get_working_date() -> str:
    return get_setting("working_date", date.today().isoformat())


def set_working_date(new_date: str) -> None:
    set_setting("working_date", new_date)


def get_secret_word() -> str:
    return get_setting("secret_word", config.SECRET_WORD)


def set_secret_word(new_word: str) -> None:
    set_setting("secret_word", new_word)


def verify_secret_word(word: str) -> bool:
    return word.strip() == get_secret_word()


# --- Undo ---

def get_last_undoable_action() -> Optional[Dict[str, Any]]:
    with get_connection() as conn:
        cursor = conn.execute(
            "SELECT * FROM undo_log WHERE reversed = 0 ORDER BY id DESC LIMIT 1"
        )
        row = cursor.fetchone()
        return dict(row) if row else None


def undo_last_action(performed_by: str = "") -> Tuple[bool, str]:
    """
    Undo the last stock-affecting action.
    Restores current_stock AND decrements the relevant sub-column so the
    daily report accurately reflects the reversal.
    """
    last = get_last_undoable_action()
    if not last:
        return False, "No action to undo."

    item_id      = last["item_id"]
    action       = last["action"]
    stock_before = last["stock_before"]
    quantity     = abs(last["quantity"])

    item = get_item_by_id(item_id)
    if not item:
        return False, f"Item '{item_id}' no longer exists."

    current_stock = item["current_stock"]
    unit          = item["unit"]
    item_name     = item["item_name"]

    try:
        with get_connection() as conn:
            now = datetime.now().isoformat()

            if action in _SUBTRACT_ACTIONS:
                new_stock = current_stock + quantity
                sub_col = _ACTION_COLUMN_MAP.get(action)
                if sub_col:
                    conn.execute(
                        f"UPDATE items SET current_stock = ?,"
                        f" {sub_col} = MAX(0, {sub_col} - ?),"
                        f" last_updated = ? WHERE item_id = ?",
                        (new_stock, quantity, now, item_id)
                    )
                else:
                    conn.execute(
                        "UPDATE items SET current_stock = ?, last_updated = ? WHERE item_id = ?",
                        (new_stock, now, item_id)
                    )

            elif action in _ADD_ACTIONS:
                new_stock = max(0.0, current_stock - quantity)
                sub_col = _ACTION_COLUMN_MAP.get(action)
                if sub_col:
                    conn.execute(
                        f"UPDATE items SET current_stock = ?,"
                        f" {sub_col} = MAX(0, {sub_col} - ?),"
                        f" last_updated = ? WHERE item_id = ?",
                        (new_stock, quantity, now, item_id)
                    )
                else:
                    conn.execute(
                        "UPDATE items SET current_stock = ?, last_updated = ? WHERE item_id = ?",
                        (new_stock, now, item_id)
                    )

            elif action == "stock adjustment":
                new_stock = stock_before   # restore to pre-adjustment value
                conn.execute(
                    "UPDATE items SET current_stock = ?, last_updated = ? WHERE item_id = ?",
                    (new_stock, now, item_id)
                )

            else:
                return False, f"Cannot undo action type: '{action}'."

            # Record the reversal in the transaction log
            conn.execute(
                """INSERT INTO transactions (
                    timestamp, date, item_id, item_name, action, quantity, unit,
                    stock_before, stock_after, performed_by, notes
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    now, date.today().isoformat(), item_id, item_name,
                    f"undo:{action}", quantity, unit,
                    current_stock, new_stock, performed_by,
                    f"Undo of {action}: restored from {current_stock} to {new_stock}"
                )
            )

            conn.execute("UPDATE undo_log SET reversed = 1 WHERE id = ?", (last["id"],))

        return (
            True,
            f"Undo successful: {item_name} restored from {current_stock} to {new_stock} {unit}.",
        )
    except Exception as e:
        return False, f"Error during undo: {str(e)}"


# --- Backup & Restore ---

def create_backup(note: str = "") -> Tuple[bool, str, Optional[str]]:
    try:
        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = Path(config.DATABASE_PATH).parent / "backups"
        backup_dir.mkdir(exist_ok=True)

        backup_filename = f"prometheus_backup_{timestamp}.db"
        backup_path = backup_dir / backup_filename

        shutil.copy2(config.DATABASE_PATH, backup_path)

        with get_connection() as conn:
            conn.execute(
                "INSERT INTO backups (filename, created_at, note) VALUES (?, ?, ?)",
                (str(backup_path), datetime.now().isoformat(), note)
            )

        return True, f"Backup created: {backup_filename}", str(backup_path)
    except Exception as e:
        return False, f"Backup failed: {str(e)}", None


def list_backups() -> List[Dict[str, Any]]:
    with get_connection() as conn:
        return [
            dict(row)
            for row in conn.execute(
                "SELECT * FROM backups ORDER BY created_at DESC"
            ).fetchall()
        ]


def restore_backup(backup_path: str) -> Tuple[bool, str]:
    try:
        if not os.path.exists(backup_path):
            return False, f"Backup file not found: {backup_path}"

        safety_ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        safety_path = (
            Path(config.DATABASE_PATH).parent
            / "backups"
            / f"prometheus_pre_restore_{safety_ts}.db"
        )
        safety_path.parent.mkdir(exist_ok=True)
        if os.path.exists(config.DATABASE_PATH):
            shutil.copy2(config.DATABASE_PATH, safety_path)

        shutil.copy2(backup_path, config.DATABASE_PATH)
        return True, "Database restored successfully."
    except Exception as e:
        return False, f"Restore failed: {str(e)}"


# --- Reset Day ---

def reset_day(performed_by: str = "") -> Tuple[bool, str]:
    """
    Carry current_stock -> starting_stock and zero all daily counters.
    """
    try:
        with get_connection() as conn:
            today = date.today().isoformat()
            now   = datetime.now().isoformat()

            for item in conn.execute("SELECT * FROM items").fetchall():
                conn.execute(
                    """UPDATE items SET
                        starting_stock    = ?,
                        used              = 0,
                        purchased         = 0,
                        damaged           = 0,
                        wipro_in          = 0,
                        wipro_out         = 0,
                        rajagiri_main     = 0,
                        woods             = 0,
                        garden_cafe       = 0,
                        bba_canteen       = 0,
                        bba_tea_counter   = 0,
                        purpose           = '',
                        working_date      = ?,
                        last_updated      = ?,
                        last_updated_by   = ?
                    WHERE item_id = ?""",
                    (item["current_stock"], today, now, performed_by, item["item_id"])
                )

            conn.execute(
                "INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)",
                ("working_date", today)
            )

        return True, f"Day reset complete. Working date: {today}. All counters reset."
    except Exception as e:
        return False, f"Reset day failed: {str(e)}"


# --- Bulk CSV Import ---

def bulk_import_items(items_data: List[Dict[str, Any]]) -> Tuple[int, int, List[str]]:
    """Import a list of item dicts. Handles both 'avg_daily_usage' and
    'average_daily_usage' column names (seed_items.csv uses the long form)."""
    success = failed = 0
    errors: List[str] = []

    for idx, row in enumerate(items_data, 1):
        item_id   = str(row.get("item_id",   "")).strip()
        item_name = str(row.get("item_name", "")).strip()
        if not item_id or not item_name:
            failed += 1
            errors.append(f"Row {idx}: Missing item_id or item_name")
            continue

        # Accept either column name for avg daily usage
        avg_raw = row.get("avg_daily_usage") or row.get("average_daily_usage") or 0

        try:
            ok, msg = add_item(
                item_id=item_id,
                item_name=item_name,
                unit=str(row.get("unit", "pcs")).strip() or "pcs",
                starting_stock=float(row.get("starting_stock", 0) or 0),
                current_stock=float(row.get("current_stock", 0) or 0),
                avg_daily_usage=float(avg_raw),
                location=str(row.get("location", "")).strip(),
                category=str(row.get("category", "")).strip(),
            )
        except (ValueError, TypeError) as e:
            failed += 1
            errors.append(f"Row {idx}: Invalid numeric value - {e}")
            continue

        if ok:
            success += 1
        else:
            failed += 1
            errors.append(f"Row {idx}: {msg}")

    return success, failed, errors


# --- CSV / Excel Export ---

def export_to_csv() -> Tuple[bool, str, Optional[str]]:
    try:
        import csv
        items = list_all_items()
        if not items:
            return False, "No items to export.", None

        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = Path(config.DATABASE_PATH).parent / "exports"
        export_dir.mkdir(exist_ok=True)
        filepath = export_dir / f"prometheus_export_{timestamp}.csv"

        fieldnames = [
            "item_id", "item_name", "unit", "starting_stock", "current_stock",
            "used", "wipro_in", "wipro_out", "rajagiri_main", "woods",
            "garden_cafe", "bba_canteen", "bba_tea_counter",
            "purchased", "damaged", "avg_daily_usage",
            "location", "category", "purpose", "last_updated",
        ]

        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for item in items:
                writer.writerow({k: item.get(k, "") for k in fieldnames})

        return True, f"Exported {len(items)} items to CSV.", str(filepath)
    except Exception as e:
        return False, f"Export failed: {str(e)}", None


def _col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter(s). Handles A-Z and AA-AZ."""
    if n <= 26:
        return chr(64 + n)
    return chr(64 + (n - 1) // 26) + chr(65 + (n - 1) % 26)


def export_to_excel() -> Tuple[bool, str, Optional[str]]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        items = list_all_items()
        if not items:
            return False, "No items to export.", None

        timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_dir = Path(config.DATABASE_PATH).parent / "exports"
        export_dir.mkdir(exist_ok=True)
        filepath   = export_dir / f"prometheus_export_{timestamp}.xlsx"

        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        header_font  = Font(bold=True, color="FFFFFF")
        header_fill  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        thin_border  = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"),  bottom=Side(style="thin"),
        )

        headers = [
            "Item ID", "Item Name", "Unit", "Starting Stock", "Current Stock",
            "Used", "Wipro In", "Wipro Out", "Rajagiri Main", "Woods",
            "Garden Cafe", "BBA Canteen", "BBA Tea Counter",
            "Purchased", "Damaged", "Avg Daily Usage",
            "Location", "Category", "Purpose", "Last Updated",
        ]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin_border

        field_keys = [
            "item_id", "item_name", "unit", "starting_stock", "current_stock",
            "used", "wipro_in", "wipro_out", "rajagiri_main", "woods",
            "garden_cafe", "bba_canteen", "bba_tea_counter",
            "purchased", "damaged", "avg_daily_usage",
            "location", "category", "purpose", "last_updated",
        ]

        for row_idx, item in enumerate(items, 2):
            for col_idx, key in enumerate(field_keys, 1):
                ws.cell(row=row_idx, column=col_idx, value=item.get(key, ""))

        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[_col_letter(col_idx)].width = 16

        wb.save(filepath)
        return True, f"Exported {len(items)} items to Excel.", str(filepath)
    except Exception as e:
        return False, f"Excel export failed: {str(e)}", None
